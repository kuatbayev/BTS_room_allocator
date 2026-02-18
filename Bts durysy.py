# -*- coding: utf-8 -*-
import pandas as pd
import random
from tkinter import Tk, Label, Button, filedialog, messagebox
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter


def build_assignments(students_df, available_rooms, max_per_room=23, max_per_class_in_room=2, attempts=400):
    class_sizes = students_df['Сыныбы'].value_counts().to_dict()
    best_assignments = None
    best_unassigned = None

    for _ in range(attempts):
        class_order = sorted(class_sizes.keys(), key=lambda cls: (-class_sizes[cls], random.random()))
        ordered_students = []

        for cls in class_order:
            cls_students = students_df[students_df['Сыныбы'] == cls].sample(frac=1).to_dict('records')
            ordered_students.extend(cls_students)

        room_assignments = {room: [] for room in available_rooms}
        room_sizes = {room: 0 for room in available_rooms}
        room_class_counts = {room: defaultdict(int) for room in available_rooms}
        unassigned_students = []

        for student in ordered_students:
            student_class = student['Сыныбы']
            candidate_rooms = [
                room for room in available_rooms
                if room_sizes[room] < max_per_room and room_class_counts[room][student_class] < max_per_class_in_room
            ]

            if not candidate_rooms:
                unassigned_students.append(student)
                continue

            min_class_count = min(room_class_counts[room][student_class] for room in candidate_rooms)
            candidate_rooms = [room for room in candidate_rooms if room_class_counts[room][student_class] == min_class_count]

            min_room_size = min(room_sizes[room] for room in candidate_rooms)
            candidate_rooms = [room for room in candidate_rooms if room_sizes[room] == min_room_size]

            chosen_room = random.choice(candidate_rooms)
            room_assignments[chosen_room].append(student)
            room_sizes[chosen_room] += 1
            room_class_counts[chosen_room][student_class] += 1

        if best_unassigned is None or len(unassigned_students) < len(best_unassigned):
            best_assignments = room_assignments
            best_unassigned = unassigned_students

        if not unassigned_students:
            break

    return best_assignments, best_unassigned


def assign_students_to_rooms(rooms_file_path, students_file_path):
    try:
        rooms_df = pd.read_excel(rooms_file_path, engine='openpyxl')
        available_rooms = rooms_df['Кабинет'].dropna().tolist()

        df = pd.read_excel(students_file_path, engine='openpyxl', dtype={'ИИН': str})

        required_cols = {'ИИН', 'Сыныбы', 'Тегі', 'Аты'}
        if not required_cols.issubset(set(df.columns)):
            messagebox.showerror(
                'Қате',
                'Excel файлы келесі бағандарды қамтуы керек: ИИН, Сыныбы, Тегі, Аты'
            )
            return

        if 'Кабинеттер' in df.columns:
            df = df.drop(columns=['Кабинеттер'])

        room_assignments, unassigned_students = build_assignments(df, available_rooms)
        for student in unassigned_students:
            print(f"⚠️ Орналастыру мүмкін болмады: {student['Аты']} {student['Тегі']} ({student['Сыныбы']})")

        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
        output_file_ready = f'Дайын_тізім_{timestamp}.xlsx'
        output_file_reference = f'Анықтамаға_іліп_қою_үшін_{timestamp}.xlsx'

        with pd.ExcelWriter(output_file_ready, engine='openpyxl') as writer:
            for room in available_rooms:
                students = room_assignments.get(room, [])
                df_room = pd.DataFrame(students)
                if df_room.empty:
                    df_room = pd.DataFrame(columns=['ИИН', 'Сыныбы', 'Тегі', 'Аты'])
                if 'Сыныбы' in df_room.columns:
                    df_room = df_room.sort_values(by='Сыныбы').reset_index(drop=True)
                df_room.insert(0, '№', range(1, len(df_room) + 1))
                df_room.to_excel(writer, sheet_name=room, startrow=1, index=False)

        with pd.ExcelWriter(output_file_reference, engine='openpyxl') as writer:
            for room in available_rooms:
                students = room_assignments.get(room, [])
                df_room = pd.DataFrame(students)
                if df_room.empty:
                    df_room = pd.DataFrame(columns=['ИИН', 'Сыныбы', 'Тегі', 'Аты'])
                if 'ИИН' in df_room.columns:
                    df_room = df_room.drop(columns=['ИИН'])
                if len(df_room.columns) >= 5:
                    df_room = df_room.drop(columns=[df_room.columns[4]])
                if 'Сыныбы' in df_room.columns:
                    df_room = df_room.sort_values(by='Сыныбы').reset_index(drop=True)
                df_room.insert(0, '№', range(1, len(df_room) + 1))
                df_room.to_excel(writer, sheet_name=room, startrow=1, index=False)

        for output_file in [output_file_ready, output_file_reference]:
            wb = load_workbook(output_file)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.merge_cells('A1:E1')
                ws['A1'] = f'Кабинет: {sheet}'
                ws['A1'].font = Font(name='Times New Roman', size=28, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row > 1:
                            cell.border = thin_border
                            cell.font = Font(name='Times New Roman', size=14)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 5

            wb.save(output_file)

        if unassigned_students:
            unassigned_df = pd.DataFrame(unassigned_students)[['ИИН', 'Сыныбы', 'Тегі', 'Аты']]
            unassigned_file = f'Орналастыру_мүмкін_болмады_{timestamp}.xlsx'
            unassigned_df.to_excel(unassigned_file, index=False)

        messagebox.showinfo(
            '✅ Дайын',
            f'Файлдар сәтті сақталды:\n{output_file_ready}\n{output_file_reference}'
        )

    except Exception as e:
        messagebox.showerror('Қате', f'Қате орын алды:\n{str(e)}')


def choose_files_and_assign():
    rooms_file_path = filedialog.askopenfilename(
        title='Кабинет файлын таңдаңыз',
        filetypes=[('Excel файлдары', '*.xlsx *.xls')]
    )
    if not rooms_file_path:
        return

    students_file_path = filedialog.askopenfilename(
        title='Оқушылар файлын таңдаңыз',
        filetypes=[('Excel файлдары', '*.xlsx *.xls')]
    )
    if not students_file_path:
        return

    assign_students_to_rooms(rooms_file_path, students_file_path)


root = Tk()
root.title('Оқушыларды кабинеттерге бөлу')
root.geometry('400x200')

Label(root, text='Оқушыларды кабинетке бөлу жүйесі', font=('Times New Roman', 16)).pack(pady=20)
Button(root, text='📂 Excel файлын таңдау', font=('Times New Roman', 14), command=choose_files_and_assign).pack(pady=10)

root.mainloop()

