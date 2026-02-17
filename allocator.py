import random
from collections import defaultdict
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter, column_index_from_string


REQUIRED_STUDENT_COLUMNS = {'ИИН', 'Сыныбы', 'Тегі', 'Аты'}


def validate_inputs(rooms_df: pd.DataFrame, students_df: pd.DataFrame):
    if 'Кабинет' not in rooms_df.columns:
        raise ValueError("В файле кабинетов не найден столбец 'Кабинет'.")

    missing = REQUIRED_STUDENT_COLUMNS - set(students_df.columns)
    if missing:
        missing_cols = ', '.join(sorted(missing))
        raise ValueError(f"В файле учеников отсутствуют столбцы: {missing_cols}")


def build_assignments(students_df, available_rooms, max_per_room=23, max_per_class_in_room=3, attempts=400):
    class_sizes = students_df['Сыныбы'].value_counts().to_dict()
    best_assignments = None
    best_unassigned = None

    for _ in range(attempts):
        class_order = sorted(class_sizes.keys(), key=lambda cls: (-class_sizes[cls], random.random()))
        ordered_students = []

        for cls in class_order:
            cls_students = students_df[students_df['Сыныбы'] == cls].sample(frac=1).to_dict('records')
            ordered_students.extend(cls_students)

        room_assignments = {room: [] for room in available_rooms}
        room_sizes = {room: 0 for room in available_rooms}
        room_class_counts = {room: defaultdict(int) for room in available_rooms}
        unassigned_students = []

        for student in ordered_students:
            student_class = student['Сыныбы']
            candidate_rooms = [
                room for room in available_rooms
                if room_sizes[room] < max_per_room and room_class_counts[room][student_class] < max_per_class_in_room
            ]

            if not candidate_rooms:
                unassigned_students.append(student)
                continue

            min_class_count = min(room_class_counts[room][student_class] for room in candidate_rooms)
            candidate_rooms = [room for room in candidate_rooms if room_class_counts[room][student_class] == min_class_count]

            min_room_size = min(room_sizes[room] for room in candidate_rooms)
            candidate_rooms = [room for room in candidate_rooms if room_sizes[room] == min_room_size]

            chosen_room = random.choice(candidate_rooms)
            room_assignments[chosen_room].append(student)
            room_sizes[chosen_room] += 1
            room_class_counts[chosen_room][student_class] += 1

        if best_unassigned is None or len(unassigned_students) < len(best_unassigned):
            best_assignments = room_assignments
            best_unassigned = unassigned_students

        if not unassigned_students:
            break

    return best_assignments, best_unassigned


def _format_workbook(workbook_path: str, title_last_col: str = 'E'):
    wb = load_workbook(workbook_path)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        expected_max_col = column_index_from_string(title_last_col)

        # Remove accidental empty columns to the right (e.g. empty E in reference file)
        if ws.max_column > expected_max_col:
            ws.delete_cols(expected_max_col + 1, ws.max_column - expected_max_col)

        ws.merge_cells(f'A1:{title_last_col}1')
        ws['A1'] = f'Кабинет: {sheet}'
        ws['A1'].font = Font(name='Times New Roman', size=28, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, expected_max_col + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name='Times New Roman', size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.value is not None and str(cell.value) != '':
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(workbook_path)


def _build_sheet_df(students):
    df_room = pd.DataFrame(students)
    if df_room.empty:
        df_room = pd.DataFrame(columns=['ИИН', 'Сыныбы', 'Тегі', 'Аты'])
    if 'Сыныбы' in df_room.columns:
        df_room = df_room.sort_values(by='Сыныбы').reset_index(drop=True)
    df_room.insert(0, '№', range(1, len(df_room) + 1))
    return df_room


def generate_outputs(
    rooms_df: pd.DataFrame,
    students_df: pd.DataFrame,
    max_per_room: int = 23,
    max_per_class_in_room: int = 3,
    attempts: int = 400,
):
    validate_inputs(rooms_df, students_df)

    available_rooms = rooms_df['Кабинет'].dropna().astype(str).tolist()
    students_df = students_df.copy()
    students_df['ИИН'] = students_df['ИИН'].astype(str)

    room_assignments, unassigned_students = build_assignments(
        students_df,
        available_rooms,
        max_per_room=max_per_room,
        max_per_class_in_room=max_per_class_in_room,
        attempts=attempts,
    )

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    ready_name = f'Дайын_тізім_{timestamp}.xlsx'
    reference_name = f'Анықтамаға_іліп_қою_үшін_{timestamp}.xlsx'

    with pd.ExcelWriter(ready_name, engine='openpyxl') as writer:
        for room in available_rooms:
            df_room = _build_sheet_df(room_assignments.get(room, []))
            df_room.to_excel(writer, sheet_name=room, startrow=1, index=False)

    with pd.ExcelWriter(reference_name, engine='openpyxl') as writer:
        for room in available_rooms:
            df_room = _build_sheet_df(room_assignments.get(room, []))
            # Reference file must contain exactly 4 columns: №, Сыныбы, Тегі, Аты
            df_room = df_room[['№', 'Сыныбы', 'Тегі', 'Аты']]
            df_room.to_excel(writer, sheet_name=room, startrow=1, index=False)

    _format_workbook(ready_name, title_last_col='E')
    _format_workbook(reference_name, title_last_col='D')

    unassigned_name = None
    if unassigned_students:
        unassigned_name = f'Орналастыру_мүмкін_болмады_{timestamp}.xlsx'
        unassigned_df = pd.DataFrame(unassigned_students)[['ИИН', 'Сыныбы', 'Тегі', 'Аты']]
        unassigned_df.to_excel(unassigned_name, index=False)

    result = {
        'ready_name': ready_name,
        'reference_name': reference_name,
        'unassigned_name': unassigned_name,
        'unassigned_count': len(unassigned_students),
        'assigned_count': len(students_df) - len(unassigned_students),
        'total_count': len(students_df),
    }
    return result


def read_excel_from_upload(uploaded_file) -> pd.DataFrame:
    content = BytesIO(uploaded_file.getvalue())
    return pd.read_excel(content, engine='openpyxl', dtype={'ИИН': str})
